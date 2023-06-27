import os
import cv2
from matplotlib import pyplot as plt
from src.OCRFunctions.OCR import applyOCR
from sklearn.metrics import precision_recall_fscore_support
from sklearn.metrics import matthews_corrcoef

import openpyxl
from PIL import Image
def FetchImagesAndGroundTruth(imagePath,outputPath, functions, tesseract_save_path, output_detected_text):
    folders = os.listdir(imagePath)
    excelfile = openpyxl.Workbook()
    excelsheet = excelfile.active
    excelsheet.title = "Results"
    for idx,function in enumerate(functions):
        excelsheet.cell(column=idx+2, row=1, value = str(function))

    for fncidx, function in enumerate(functions):
        global_img_idx = 0
        ground_truths = []
        predictions = []
        all_predictions = 0
        correct_predictions = 0
        for folder in folders:
            groundTruth = folder
            folderPath = os.path.join(imagePath, folder)
            images = os.listdir(folderPath)
            for imgidx, imagename in enumerate(images):
                global_img_idx = global_img_idx +1
                excelsheet.cell(column=1, row=global_img_idx+1, value=imagename)
                #print(os.path.join(folderPath,imagename))
                #print(os.path.exists(os.path.join(folderPath,imagename)))

                image = cv2.imread(os.path.join(folderPath,imagename))
                #plt.imshow(image)
                #plt.show()

                if function == "None":
                    processedImage = image
                    txtname = os.path.join(outputPath, str(function) + os.path.splitext(imagename)[0])
                else:
                    processedImage = function(image)
                    txtname = os.path.join(outputPath, str(function.__name__) + os.path.splitext(imagename)[0])
                if output_detected_text:
                    calculatedLabel, text = applyOCR(processedImage, tesseract_save_path, output_detected_text)

                    with open(txtname + ".txt", "w", encoding ='utf-8') as file:
                        file.write(text)
                else:
                    calculatedLabel,text = applyOCR(processedImage, tesseract_save_path, output_detected_text)


                #wasCorrect, Label = determineCorrectness(calculatedLabel,groundTruth)
                #if wasCorrect:
                ground_truths.append(groundTruth)
                predictions.append(calculatedLabel)
                all_predictions = all_predictions +1
                if calculatedLabel == groundTruth:
                    correct_predictions = correct_predictions + 1
                    excelsheet.cell(column=fncidx + 2, row=global_img_idx + 1, value="Correct: " + str(calculatedLabel))
                else:
                    excelsheet.cell(column=fncidx + 2, row=global_img_idx + 1, value="Wrong Label determined. True Label: " + str(groundTruth) + " Detected_Label: " + str(calculatedLabel))
#                else:
#                    excelsheet.cell(column=fncidx + 2, row=imgidx+2, value="No Label determined")
        print("Accuracy")
        print((correct_predictions / all_predictions) * 100)
        print("f1 score")
        print(ground_truths)
        print(predictions)
        precision, recall, f1, _ = precision_recall_fscore_support(ground_truths, predictions, average='macro')
        print(f1)
        print("MMC")
        MCC = matthews_corrcoef(ground_truths, predictions)
        print(MCC)
        global_img_idx = global_img_idx +1
        excelsheet.cell(column=fncidx + 2, row=global_img_idx + 3, value=(correct_predictions / all_predictions) * 100)
        excelsheet.cell(column=fncidx + 2, row=global_img_idx + 4, value=precision)
        excelsheet.cell(column=fncidx + 2, row=global_img_idx + 5, value=recall)
        excelsheet.cell(column=fncidx + 2, row=global_img_idx + 6, value=f1)
        excelsheet.cell(column=fncidx + 2, row=global_img_idx + 7, value=MCC)
        excelsheet.cell(column=1, row=global_img_idx + 3, value="Accuracy")
        excelsheet.cell(column=1, row=global_img_idx + 4, value="precision")
        excelsheet.cell(column=1, row=global_img_idx + 5, value="recall")
        excelsheet.cell(column=1, row=global_img_idx + 6, value="f1")
        excelsheet.cell(column=1, row=global_img_idx + 7, value="MCC")

    counter = 2
    output = f'{os.path.join(outputPath,"Accuracy OCR.xlsx")}'
    while os.path.exists(output):
        output = f'{os.path.join(outputPath, "Accuracy OCR" + str(counter) + ".xlsx")}'
        counter = counter + 1
    excelfile.save(output)
