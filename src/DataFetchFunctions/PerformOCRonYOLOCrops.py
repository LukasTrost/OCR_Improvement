import os
import cv2
from matplotlib import pyplot as plt
from src.OCRFunctions.OCR import applyOCR
from sklearn.metrics import precision_recall_fscore_support
from sklearn.metrics import matthews_corrcoef
import openpyxl
def PerformOCRonYOLOCrops(imagePath,outputPath, functions, tesseract_save_path, output_detected_text):
    folders = os.listdir(imagePath)
    print(folders)
    excelfile = openpyxl.Workbook()
    excelsheet = excelfile.active
    excelsheet.title = "Results"
    for idx,function in enumerate(functions):
        excelsheet.cell(column=idx+2, row=1, value = str(function))

    for fncidx, function in enumerate(functions):
        global_img_idx = 0
        ground_truths = []
        predictions = []
        all_predictions = 0
        correct_predictions = 0
        for folder_idx, folder in enumerate(folders):
            global_img_idx = global_img_idx + 1
            excelsheet.cell(column=1, row=global_img_idx + 1, value=folder)
            groundTruth = folder.split("_test")[0]
            print(groundTruth)


            folder_partial_Path = os.path.join(os.path.join(os.path.join(imagePath, folder),"predict"), "crops")
            if os.path.exists(folder_partial_Path):
                folder_class = os.listdir(folder_partial_Path)[0]
                folderPath = os.path.join(folder_partial_Path,folder_class)
                image_path = os.listdir(folderPath)[0]


                image = cv2.imread(os.path.join(folderPath,image_path))

                if function == "None":
                    processedImage = image
                    txtname = os.path.join(outputPath, str(function) + os.path.splitext(image_path)[0])
                else:
                    processedImage = function(image)
                    txtname = os.path.join(outputPath, str(function.__name__) + os.path.splitext(image_path)[0])
                if output_detected_text:
                    calculatedLabel, text = applyOCR(processedImage, tesseract_save_path, output_detected_text)
                    with open(txtname + ".txt", "w") as file:
                        file.write(text)
                else:
                    calculatedLabel, text = applyOCR(processedImage, tesseract_save_path, output_detected_text)



                #wasCorrect, Label = determineCorrectness(max_entry,groundTruth)
                ground_truths.append(groundTruth)
                predictions.append(calculatedLabel)
                all_predictions = all_predictions +1
                if calculatedLabel == "None":
                    calculatedLabel = "Background"
                if calculatedLabel == groundTruth:
                    correct_predictions = correct_predictions +1
                    excelsheet.cell(column=fncidx + 2, row=global_img_idx + 1, value="Correct: " + str(calculatedLabel))
                else:
                    print("calculated: ", calculatedLabel, "truth", groundTruth)
                    excelsheet.cell(column=fncidx + 2, row=global_img_idx + 1,
                                    value="Wrong Label determined. True Label: " + str(
                                        groundTruth) + " Detected_Label: " + str(calculatedLabel))
                    #                else:
                    #                    excelsheet.cell(column=fncidx + 2, row=imgidx+2, value="No Label determined")
            else:
                all_predictions = all_predictions + 1
                predictions.append("Background")
                ground_truths.append(groundTruth)

                if groundTruth == "Background":
                    print("groundTruth is ", groundTruth, " and not Background")
                    correct_predictions = correct_predictions +1
                    excelsheet.cell(column=fncidx + 2, row=global_img_idx + 1, value="Correct: " + "Background")
                else:
                    print("groundTruth is ", groundTruth, " and not Background")
                    excelsheet.cell(column=fncidx + 2, row=global_img_idx + 1,
                                    value="Wrong Label determined. True Label: " + str(
                                        groundTruth) + " Detected_Label: " + "Background")


        print("Accuracy")
        print((correct_predictions / all_predictions) * 100)
        print("f1 score")
        print(ground_truths)
        print(predictions)
        precision, recall, f1, _ = precision_recall_fscore_support(ground_truths, predictions, average='macro')
        print(f1)
        print("MMC")
        MCC = matthews_corrcoef(ground_truths, predictions)
        print(MCC)
        global_img_idx = global_img_idx + 1
        excelsheet.cell(column=fncidx + 2, row=global_img_idx + 3,
                        value=(correct_predictions / all_predictions) * 100)
        excelsheet.cell(column=fncidx + 2, row=global_img_idx + 4, value=precision)
        excelsheet.cell(column=fncidx + 2, row=global_img_idx + 5, value=recall)
        excelsheet.cell(column=fncidx + 2, row=global_img_idx + 6, value=f1)
        excelsheet.cell(column=fncidx + 2, row=global_img_idx + 7, value=MCC)
        excelsheet.cell(column=1, row=global_img_idx + 3, value="Accuracy")
        excelsheet.cell(column=1, row=global_img_idx + 4, value="precision")
        excelsheet.cell(column=1, row=global_img_idx + 5, value="recall")
        excelsheet.cell(column=1, row=global_img_idx + 6, value="f1")
        excelsheet.cell(column=1, row=global_img_idx + 7, value="MCC")
    output = f'{os.path.join(outputPath,"Accuracy OCR.xlsx")}'
    counter = 0
    while os.path.exists(output):
        output = f'{os.path.join(outputPath, "Accuracy OCR" + str(counter) + ".xlsx")}'
        counter = counter + 1
    print("saving now")
    excelfile.save(output)
